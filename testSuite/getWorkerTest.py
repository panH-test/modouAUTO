"""
工种code及name导出到excel
time：2021-11-29
author：hanpan
"""


import openpyxl
import requests
import json

mywb=openpyxl.Workbook()
mywb.create_sheet(index=0,title='worker')

url='http://120.92.35.79:8083/dict/workType'
res=requests.get(url);
res.encoding="utf-8"

jst=json.loads(res.text)
workercode=jst['data']




def ExcelWrite(result):
    mywb = openpyxl.Workbook()
    sheet=mywb.active;
    row=1;
    for i in result:
        col=64;
        row+=1;
        for j in i:
            col+=1;
            sheet[chr(col)+str(row)]=i[j]
    mywb.save('C:\\Users\86133\Desktop\gongzhong.xlsx')

ExcelWrite(workercode);